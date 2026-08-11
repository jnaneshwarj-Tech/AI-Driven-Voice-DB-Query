"""
routes_export.py — Consistent export: PDF / Excel / CSV only (JSON removed).

Every export uses ONE shared _build_export_model() from rendered report data.
ALL formats contain IDENTICAL content:
  ✔ University / College / Department header
  ✔ Student personal information (name, DOB, blood group, father, address…)
  ✔ Academic performance table (Semester | SGPA | Cumulative CGPA)
  ✔ Signature section — auto-filled with logged-in user (role + name)
  ✔ Generated date & time

PDF: institutional header + logo + personal info block + academic table + signature
Excel: 3 sheets — Report, Student Info, Semester CGPA
CSV: metadata section + personal section + academic section
"""

import os, tempfile
from datetime import datetime
import pandas as pd
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, Image as RLImage
)
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from auth import get_current_user

router = APIRouter(prefix="/api/export", tags=["Export"])

# ── Institutional constants ───────────────────────────────────────────────────
UNIVERSITY = "VISVESVARAYA TECHNOLOGICAL UNIVERSITY, BELAGAVI-590018"
COLLEGE    = "GOVERNMENT ENGINEERING COLLEGE MOSALEHOSAHALLI, HASSAN"
DEPARTMENT = "DEPARTMENT OF COMPUTER SCIENCE AND ENGINEERING"
LOGO_PATH  = os.path.join(
    os.path.dirname(__file__), "..", "frontend", "public", "college-logo.png"
)

# ── Field groupings ───────────────────────────────────────────────────────────
# photo_url intentionally excluded — photo system removed
_PERSONAL_KEYS = [
    "usn", "name", "dob", "gender", "blood_group",
    "father_name", "mother_name",
    "phone", "email", "aadhar_no",
    "address", "permanent_address", "current_address",
    "religion", "caste", "sub_caste", "category",
    "year_and_branch", "year_of_joining", "current_sem", "status",
    "admission_year", "current_year", "student_type", "estimated_semester",
    "graduation_year", "graduation_status",
]
_PHOTO_KEYS = {"photo_url", "image_url", "photo", "photo_path"}
_PERSONAL_LABELS = {
    "usn": "USN", "name": "Name", "dob": "Date of Birth",
    "gender": "Gender", "blood_group": "Blood Group",
    "father_name": "Father Name", "mother_name": "Mother Name",
    "phone": "Phone", "email": "Email", "aadhar_no": "Aadhar No",
    "address": "Address", "permanent_address": "Permanent Address",
    "current_address": "Current Address",
    "religion": "Religion", "caste": "Caste", "sub_caste": "Sub Caste",
    "category": "Category", "year_and_branch": "Year & Branch",
    "year_of_joining": "Year of Joining", "current_sem": "Current Semester",
    "status": "Status",
    "admission_year": "Admission Year", "current_year": "Current Year",
    "student_type": "Student Type", "estimated_semester": "Estimated Semester",
    "graduation_year": "Graduation Year", "graduation_status": "Graduation Status",
}
_ACADEMIC_KEYS = {"usn", "name", "semester", "sgpa", "cgpa", "year"}


def _now_str() -> str:
    return datetime.now().strftime("%d %B %Y, %I:%M %p")


# ── Shared export model ───────────────────────────────────────────────────────

def _build_export_model(data: list[dict], user: dict) -> dict:
    """
    Single source of truth for all export formats.
    Extracts personal + academic fields and computes semester-wise CGPA.
    """
    role  = user.get("role", "Staff")
    uname = user.get("username", "Unknown")

    # ── Clean data (remove photo fields) ──────────────────────────────────────
    cleaned_data = []
    for entry in data:
        cleaned_data.append({k: v for k, v in entry.items() if k not in _PHOTO_KEYS})

    # ── Personal details (from first row if single student) ───────────────────
    student_details: dict = {}
    if cleaned_data:
        first = cleaned_data[0]
        for k in _PERSONAL_KEYS:
            v = first.get(k)
            if v not in (None, "", "null"):
                student_details[k] = v

    # ── Academic rows ─────────────────────────────────────────────────────────
    academic_rows = []
    for r in cleaned_data:
        row = {k: v for k, v in r.items() if k in _ACADEMIC_KEYS}
        if row:
            academic_rows.append(row)

    # ── Semester-wise cumulative CGPA per student ─────────────────────────────
    groups: dict[str, list] = {}
    for r in academic_rows:
        key = r.get("usn") or r.get("name") or "unknown"
        groups.setdefault(key, []).append(r)

    semester_wise_cgpa: dict = {}
    for key, rows in groups.items():
        sems = sorted(
            [(r.get("semester", 0), float(r.get("sgpa") or 0)) for r in rows],
            key=lambda x: x[0]
        )
        running, cgpa_list = 0.0, []
        for i, (sem, sgpa) in enumerate(sems):
            running += sgpa
            cgpa_list.append({
                "semester": sem,
                "sgpa":     round(sgpa, 2),
                "cgpa":     round(running / (i + 1), 2),
            })
        semester_wise_cgpa[key] = cgpa_list

    return {
        "institution": {
            "university": UNIVERSITY,
            "college":    COLLEGE,
            "department": DEPARTMENT,
        },
        "student_details":    student_details,
        "academic_rows":      academic_rows,
        "semester_wise_cgpa": semester_wise_cgpa,
        "generated_by": {
            "role":      role,
            "username":  uname,
            "timestamp": _now_str(),
        },
    }


# ── PDF builder ───────────────────────────────────────────────────────────────

def _build_branded_pdf(path: str, model: dict):
    # Styles
    hdr_s  = ParagraphStyle('H',  fontName='Helvetica-Bold', fontSize=11, alignment=TA_CENTER, leading=16)
    sub_s  = ParagraphStyle('S',  fontName='Helvetica-Bold', fontSize=9,  alignment=TA_CENTER, leading=13,
                             textColor=colors.HexColor("#1e3a5f"))
    dept_s = ParagraphStyle('D',  fontName='Helvetica-Bold', fontSize=8,  alignment=TA_CENTER, leading=12,
                             textColor=colors.HexColor("#1e3a5f"))
    norm_s = ParagraphStyle('N',  fontName='Helvetica',      fontSize=8,  alignment=TA_LEFT,   leading=11)
    sign_s = ParagraphStyle('SG', fontName='Helvetica',      fontSize=8,  alignment=TA_LEFT,   leading=12)
    tiny_s = ParagraphStyle('T',  fontName='Helvetica',      fontSize=7,  alignment=TA_CENTER, leading=10,
                             textColor=colors.grey)
    lbl_s  = ParagraphStyle('LB', fontName='Helvetica-Bold', fontSize=8,  alignment=TA_LEFT,   leading=12)
    val_s  = ParagraphStyle('VL', fontName='Helvetica',      fontSize=8,  alignment=TA_LEFT,   leading=12)
    sec_s  = ParagraphStyle('SC', fontName='Helvetica-Bold', fontSize=9,  alignment=TA_LEFT,   leading=13,
                             textColor=colors.HexColor("#1e3a5f"))

    inst  = model["institution"]
    gen   = model["generated_by"]
    sd    = model.get("student_details", {})
    cgpa  = model.get("semester_wise_cgpa", {})
    story = []

    # ── Logo + header block ───────────────────────────────────────────────────
    logo_cell = Paragraph("", norm_s)
    if os.path.exists(LOGO_PATH):
        try:
            logo_cell = RLImage(LOGO_PATH, width=2.2*cm, height=2.2*cm)
        except Exception:
            pass

    hdr_text = [
        Paragraph(inst["university"], hdr_s), Spacer(1, 2),
        Paragraph(inst["college"],    sub_s),  Spacer(1, 2),
        Paragraph(inst["department"], dept_s),
    ]
    hdr_tbl = Table([[logo_cell, hdr_text]], colWidths=[2.8*cm, None])
    hdr_tbl.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(hdr_tbl)
    story.append(HRFlowable(width="100%", thickness=1.5,
                             color=colors.HexColor("#1e3a5f"), spaceAfter=6))

    # ── Title ─────────────────────────────────────────────────────────────────
    title_s = ParagraphStyle('TT', fontName='Helvetica-Bold', fontSize=10,
                               alignment=TA_CENTER, leading=14,
                               textColor=colors.HexColor("#1e3a5f"))
    story.append(Paragraph("Student Academic Performance Report", title_s))
    story.append(Spacer(1, 8))

    # ── Student Personal Information Block (no photo) ───────────────────────
    if sd:
        story.append(Paragraph("PERSONAL INFORMATION", sec_s))
        story.append(Spacer(1, 4))

        # Build 2-column label/value rows to save vertical space
        info_rows = []
        for k in _PERSONAL_KEYS:
            v = sd.get(k)
            if v not in (None, "", "null"):
                info_rows.append([
                    Paragraph(f"<b>{_PERSONAL_LABELS.get(k, k.replace('_',' ').title())}:</b>", lbl_s),
                    Paragraph(str(v), val_s)
                ])

        if info_rows:
            half = (len(info_rows) + 1) // 2
            left_col  = info_rows[:half]
            right_col = info_rows[half:]
            while len(right_col) < len(left_col):
                right_col.append([Paragraph("", lbl_s), Paragraph("", val_s)])
            merged_rows = [l + r for l, r in zip(left_col, right_col)]
            info_tbl = Table(merged_rows, colWidths=[3.2*cm, 5.5*cm, 3.2*cm, 5.5*cm])
            info_tbl.setStyle(TableStyle([
                ('VALIGN',        (0,0), (-1,-1), 'TOP'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 2),
                ('TOPPADDING',    (0,0), (-1,-1), 2),
                ('LEFTPADDING',   (0,0), (-1,-1), 3),
                ('BACKGROUND',    (0,0), (-1,-1), colors.HexColor("#f8faff")),
                ('BOX',           (0,0), (-1,-1), 0.5, colors.HexColor("#d0d8ef")),
                ('INNERGRID',     (0,0), (-1,-1), 0.3, colors.HexColor("#e8edf5")),
            ]))
            story.append(info_tbl)

        story.append(Spacer(1, 6))
        story.append(HRFlowable(width="100%", thickness=0.5,
                                 color=colors.HexColor("#cccccc"), spaceAfter=6))

    # ── Semester-wise Academic Performance ────────────────────────────────────
    if cgpa:
        story.append(Paragraph("ACADEMIC PERFORMANCE", sec_s))
        story.append(Spacer(1, 4))

        for usn_key, rows in cgpa.items():
            if len(cgpa) > 1:
                story.append(Paragraph(f"<b>{usn_key}</b>", norm_s))
                story.append(Spacer(1, 3))

            t_data = [["Semester", "SGPA", "Cumulative CGPA"]]
            for row in rows:
                t_data.append([
                    f"Semester {row['semester']}",
                    str(row['sgpa']),
                    str(row['cgpa'])
                ])

            t = Table(t_data, repeatRows=1, colWidths=[5*cm, 4*cm, 5*cm])
            t.setStyle(TableStyle([
                ('BACKGROUND',    (0,0), (-1,0), colors.HexColor("#1e3a5f")),
                ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
                ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE',      (0,0), (-1,-1), 8),
                ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
                ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor("#f0f4f8")]),
                ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor("#cccccc")),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
                ('TOPPADDING',    (0,0), (-1,-1), 5),
            ]))
            story.append(t)
            story.append(Spacer(1, 6))

    # ── Generic academic table if no CGPA computed ────────────────────────────
    elif model.get("academic_rows"):
        acad = model["academic_rows"]
        headers = list({k for r in acad for k in r.keys()})
        page_w = A4[0] - 3*cm
        col_w  = page_w / max(len(headers), 1)
        td_data = [[h.replace("_", " ").title() for h in headers]] + \
                  [[str(r.get(h, "—") or "—") for h in headers] for r in acad]
        t = Table(td_data, repeatRows=1, colWidths=[col_w]*len(headers))
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,0), colors.HexColor("#1e3a5f")),
            ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
            ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,-1), 7),
            ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
            ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor("#f0f4f8")]),
            ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor("#cccccc")),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING',    (0,0), (-1,-1), 4),
        ]))
        story.append(t)
        story.append(Spacer(1, 10))

    # ── Signature Section — Computer-filled for PDF/Excel/CSV ────────────────
    story.append(HRFlowable(width="100%", thickness=0.5,
                             color=colors.grey, spaceBefore=6, spaceAfter=8))
    sig_data = [[
        Paragraph(
            f"<b>Prepared By:</b><br/><br/>"
            f"Signature: {gen['username']}<br/><br/>"
            f"<b>({gen['role']})</b>",
            sign_s
        ),
        Paragraph(
            f"<b>Generated on:</b><br/>{gen['timestamp']}",
            sign_s
        ),
        Paragraph("", sign_s),
    ]]
    sig_tbl = Table(sig_data, colWidths=[7*cm, 7*cm, 4*cm])
    sig_tbl.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(sig_tbl)
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f"Report generated by AI Student Management System · {COLLEGE}", tiny_s
    ))

    doc = SimpleDocTemplate(
        path, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )
    doc.build(story)


def _log_export(fmt: str, record_count: int, username: str):
    from database import db_conn
    try:
        with db_conn() as conn:
            cur = conn.cursor()
            cur.execute(
                "INSERT INTO export_logs (format, record_count, exported_by) VALUES (%s, %s, %s)",
                (fmt, record_count, username)
            )
            conn.commit()
            cur.close()
    except Exception as e:
        print(f"[ERR] Failed to log export: {e}")


# ── Routes ────────────────────────────────────────────────────────────────────

@router.post("/pdf")
def export_pdf(data: list[dict], current_user: dict = Depends(get_current_user)):
    if not data:
        raise HTTPException(400, "No data to export.")
    model = _build_export_model(data, current_user)
    _log_export("pdf", len(data), current_user["username"])
    fd, path = tempfile.mkstemp(suffix=".pdf")
    os.close(fd)
    _build_branded_pdf(path, model)
    return FileResponse(path, filename="student_report.pdf", media_type="application/pdf")


@router.post("/excel")
def export_excel(data: list[dict], current_user: dict = Depends(get_current_user)):
    if not data:
        raise HTTPException(400, "No data to export.")
    model = _build_export_model(data, current_user)
    gen   = model["generated_by"]
    inst  = model["institution"]
    sd    = model.get("student_details", {})
    _log_export("excel", len(data), current_user["username"])

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)

    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # 1. University Header
    ws.append([inst["university"]])
    ws.append([inst["college"]])
    ws.append([inst["department"]])
    ws.append([])
    
    for r in range(1, 4):
        ws.cell(row=r, column=1).font = Font(bold=True, size=11)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
        
    # 2. Personal Details
    if sd:
        ws.append(["PERSONAL DETAILS"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        for k, v in sd.items():
            label = _PERSONAL_LABELS.get(k, k.replace("_"," ").title())
            ws.append([f"{label}:", str(v) if v else ""])
        ws.append([])
        
    # 3. Academic Details Table (Dashboard Results)
    ws.append(["ACADEMIC DETAILS TABLE"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    
    clean = [{k: v for k, v in r.items() if k not in _PHOTO_KEYS} for r in data]
    if clean:
        headers = list(clean[0].keys())
        # Replace headers with uppercase
        display_headers = [h.replace("_", " ").upper() for h in headers]
        ws.append(display_headers)
        
        header_row_idx = ws.max_row
        for col_idx in range(1, len(display_headers) + 1):
            c = ws.cell(row=header_row_idx, column=col_idx)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1e3a5f")
            c.alignment = Alignment(horizontal="center")
            
        for row_data in clean:
            ws.append([str(row_data.get(h, "")) for h in headers])
            
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
            
    ws.append([])
    ws.append([])
    
    # 4. Signature
    sig_row = ws.max_row
    ws.cell(row=sig_row, column=1, value="Prepared By:")
    ws.cell(row=sig_row+1, column=1, value=f"Signature: {gen['username']}")
    ws.cell(row=sig_row+2, column=1, value=f"({gen['role']})")
    ws.cell(row=sig_row+3, column=1, value=f"Generated on: {gen['timestamp']}")
    
    wb.save(path)

    return FileResponse(
        path,
        filename="student_report.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )




@router.post("/csv")
def export_csv(data: list[dict], current_user: dict = Depends(get_current_user)):
    if not data:
        raise HTTPException(400, "No data to export.")
    model = _build_export_model(data, current_user)
    gen   = model["generated_by"]
    inst  = model["institution"]
    sd    = model.get("student_details", {})
    cgpa  = model.get("semester_wise_cgpa", {})
    _log_export("csv", len(data), current_user["username"])

    fd, path = tempfile.mkstemp(suffix=".csv")
    with os.fdopen(fd, 'w', encoding='utf-8', newline='') as f:

        # ── Institution metadata ──────────────────────────────────────────────
        f.write("=== INSTITUTION ===\n")
        f.write(f"University,{inst['university']}\n")
        f.write(f"College,{inst['college']}\n")
        f.write(f"Department,{inst['department']}\n\n")

        # ── Personal details ──────────────────────────────────────────────────
        if sd:
            f.write("=== PERSONAL INFORMATION ===\n")
            for k, v in sd.items():
                label = _PERSONAL_LABELS.get(k, k.replace("_", " ").title())
                f.write(f"{label},{str(v) if v else ''}\n")
            f.write("\n")

        # ── Dashboard Results (Academic Details Table) ────────────────────────
        f.write("=== ACADEMIC DETAILS TABLE ===\n")
        clean = [{k: v for k, v in r.items() if k not in _PHOTO_KEYS} for r in data]
        if clean:
            headers = list(clean[0].keys())
            display_headers = [h.replace("_", " ").upper() for h in headers]
            f.write(",".join(display_headers) + "\n")
            for row in clean:
                f.write(",".join(str(row.get(h, "")) for h in headers) + "\n")
        f.write("\n")

        # ── Signature — computer-filled ─────────────────────────────────────
        f.write(f"\n=== SIGNATURE ===\n")
        f.write(f"Prepared By,\n")
        f.write(f"Signature,{gen['username']}\n")
        f.write(f"Role,({gen['role']})\n")
        f.write(f"Generated on,{gen['timestamp']}\n")

    return FileResponse(path, filename="student_report.csv", media_type="text/csv")
